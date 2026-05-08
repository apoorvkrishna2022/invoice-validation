import base64
import io
import json
from difflib import SequenceMatcher
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
import streamlit as st

GOLDEN_DIR = Path(__file__).parent / "golden_data"
TRIAL_DIR = Path(__file__).parent / "Trials" / "Trial_1"
PDF_DIR = Path(__file__).parent / "pdfs"
TOTAL_INVOICES = 56
FUZZY_THRESHOLD = 0.90

# ── Colours ────────────────────────────────────────────────────────────────────
GREEN  = PatternFill("solid", fgColor="D4EDDA")
YELLOW = PatternFill("solid", fgColor="FFF9C4")
RED    = PatternFill("solid", fgColor="F8D7DA")
HEADER = PatternFill("solid", fgColor="343A40")
HEADER_FONT = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)

st.set_page_config(page_title="Invoice Validation", layout="wide")
st.title("Invoice Validation Dashboard")
st.caption("Comparing Trial_1 outputs against golden data — exact match + fuzzy ≥ 90%")


# ── Data loading ───────────────────────────────────────────────────────────────
def load_metadata(path: Path) -> dict:
    with open(path) as f:
        data = json.load(f)
    return {item["field"]: item["value"] for item in data.get("metadata", [])}


def _is_empty_or_zero(v) -> bool:
    if v == "" or v is None:
        return True
    try:
        return float(v) == 0.0
    except (ValueError, TypeError):
        return False


def compare_values(golden_val, trial_val) -> tuple[bool, float, bool]:
    """Return (is_match, similarity_0_to_1, is_exact)."""
    if _is_empty_or_zero(golden_val) and _is_empty_or_zero(trial_val):
        return True, 1.0, True

    if isinstance(golden_val, (int, float)) or isinstance(trial_val, (int, float)):
        try:
            exact = float(golden_val) == float(trial_val)
            return exact, 1.0 if exact else 0.0, exact
        except (ValueError, TypeError):
            pass

    g, t = str(golden_val).strip(), str(trial_val).strip()
    if g == t:
        return True, 1.0, True

    sim = SequenceMatcher(None, g.lower(), t.lower()).ratio()
    return sim >= FUZZY_THRESHOLD, sim, False


def pdf_iframe(pdf_path: Path) -> str:
    with open(pdf_path, "rb") as f:
        b64 = base64.b64encode(f.read()).decode()
    return (
        f'<iframe src="data:application/pdf;base64,{b64}" '
        f'width="100%" height="700px" style="border:none;border-radius:6px"></iframe>'
    )


@st.cache_data
def load_all() -> list:
    records = []
    for idx in range(1, TOTAL_INVOICES + 1):
        golden_path = GOLDEN_DIR / f"{idx}.json"
        trial_path  = TRIAL_DIR  / f"{idx}.json"

        if not golden_path.exists():
            continue

        golden_meta = load_metadata(golden_path)

        if not trial_path.exists():
            records.append({"idx": idx, "status": "not_tested", "rows": [], "golden": golden_meta})
            continue

        trial_meta = load_metadata(trial_path)
        rows = []
        for field, trial_val in trial_meta.items():
            golden_val = golden_meta.get(field, "—MISSING—")
            if golden_val == "—MISSING—":
                match, sim, exact = False, 0.0, False
            else:
                match, sim, exact = compare_values(golden_val, trial_val)
            rows.append({"Field": field, "Golden": str(golden_val),
                         "Trial": str(trial_val), "Match": match,
                         "Sim": sim, "Exact": exact})

        all_match = all(r["Match"] for r in rows)
        records.append({"idx": idx, "status": "passed" if all_match else "failed", "rows": rows})
    return records


# ── Excel export ───────────────────────────────────────────────────────────────
def build_excel(records: list) -> bytes:
    wb = openpyxl.Workbook()

    # ── Summary sheet ──────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"

    tested     = [r for r in records if r["status"] != "not_tested"]
    passed     = [r for r in tested  if r["status"] == "passed"]
    failed     = [r for r in tested  if r["status"] == "failed"]
    not_tested = [r for r in records if r["status"] == "not_tested"]
    accuracy   = len(passed) / len(tested) * 100 if tested else 0

    # Metric block
    metrics = [
        ("Total Invoices", TOTAL_INVOICES),
        ("Tested",         len(tested)),
        ("Passed",         len(passed)),
        ("Failed",         len(failed)),
        ("Not Tested",     len(not_tested)),
        ("Accuracy",       f"{accuracy:.1f}%"),
    ]
    for row_i, (label, value) in enumerate(metrics, start=1):
        ws.cell(row_i, 1, label).font = BOLD
        ws.cell(row_i, 2, value)
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 15

    # Spacer
    ws.append([])

    # Invoice table header
    hdr_row = len(metrics) + 2
    for col, title in enumerate(["Invoice", "Status", "Fields Compared", "Fields Matched", "Field Accuracy"], 1):
        cell = ws.cell(hdr_row, col, title)
        cell.fill = HEADER
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for r in records:
        if r["status"] == "not_tested":
            ws.append([f"Invoice {r['idx']}", "Not Tested", "—", "—", "—"])
        else:
            total   = len(r["rows"])
            matched = sum(1 for row in r["rows"] if row["Match"])
            pct     = f"{matched/total*100:.1f}%" if total else "—"
            data_row = ws.max_row + 1
            ws.append([f"Invoice {r['idx']}", r["status"].capitalize(), total, matched, pct])
            fill = GREEN if r["status"] == "passed" else RED
            for col in range(1, 6):
                ws.cell(data_row, col).fill = fill

    for col_letter, width in [("C", 18), ("D", 16), ("E", 14)]:
        ws.column_dimensions[col_letter].width = width

    # ── One sheet per invoice ──────────────────────────────────────────────────
    for rec in records:
        sheet_name = f"Invoice {rec['idx']}"
        ws_i = wb.create_sheet(title=sheet_name)

        if rec["status"] == "not_tested":
            ws_i.append(["Status", "Not Tested"])
            ws_i.append(["No trial output for this invoice."])
            continue

        # Header row
        for col, title in enumerate(["Field", "Golden", "Trial", "Match", "Similarity %"], 1):
            cell = ws_i.cell(1, col, title)
            cell.fill = HEADER
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")

        for row_i, r in enumerate(rec["rows"], start=2):
            if r["Match"] and r["Exact"]:
                status_label, fill = "Exact", GREEN
            elif r["Match"]:
                status_label, fill = "Fuzzy", YELLOW
            else:
                status_label, fill = "Mismatch", RED

            ws_i.cell(row_i, 1, r["Field"])
            ws_i.cell(row_i, 2, r["Golden"])
            ws_i.cell(row_i, 3, r["Trial"])
            ws_i.cell(row_i, 4, status_label)
            ws_i.cell(row_i, 5, round(r["Sim"] * 100, 1))
            for col in range(1, 6):
                ws_i.cell(row_i, col).fill = fill

        ws_i.column_dimensions["A"].width = 28
        ws_i.column_dimensions["B"].width = 40
        ws_i.column_dimensions["C"].width = 40
        ws_i.column_dimensions["D"].width = 12
        ws_i.column_dimensions["E"].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── UI ─────────────────────────────────────────────────────────────────────────
records = load_all()

tested     = [r for r in records if r["status"] != "not_tested"]
passed     = [r for r in tested  if r["status"] == "passed"]
failed     = [r for r in tested  if r["status"] == "failed"]
not_tested = [r for r in records if r["status"] == "not_tested"]
accuracy   = len(passed) / len(tested) * 100 if tested else 0

c1, c2, c3, c4, c5 = st.columns(5)
c1.metric("Total Invoices", TOTAL_INVOICES)
c2.metric("Tested",  len(tested))
c3.metric("Passed",  len(passed))
c4.metric("Failed",  len(failed))
c5.metric("Accuracy", f"{accuracy:.1f}%")

st.download_button(
    label="⬇️ Download Excel Report",
    data=build_excel(records),
    file_name="invoice_validation.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)

st.divider()

STATUS_EMOJI = {"passed": "✅", "failed": "❌", "not_tested": "⬜"}

with st.sidebar:
    st.header("Invoices")
    filter_opt = st.radio("Filter", ["All", "Failed", "Not Tested"], horizontal=True)

    def filter_fn(r):
        if filter_opt == "All":
            return True
        if filter_opt == "Failed":
            return r["status"] == "failed"
        return r["status"] == "not_tested"

    visible = [r for r in records if filter_fn(r)]
    options = [f"{STATUS_EMOJI[r['status']]} Invoice {r['idx']}" for r in visible]

    if not options:
        st.info("No invoices match this filter.")
        selected_record = None
    else:
        chosen     = st.radio("Select invoice", options, label_visibility="collapsed")
        chosen_idx = visible[options.index(chosen)]["idx"]
        selected_record = next(r for r in records if r["idx"] == chosen_idx)

# ── Detail view ────────────────────────────────────────────────────────────────
if selected_record is None:
    st.info("No invoice selected.")
else:
    rec      = selected_record
    pdf_path = PDF_DIR / f"{rec['idx']}.pdf"

    if rec["status"] == "not_tested":
        st.subheader(f"Invoice {rec['idx']} — Not Tested")
        st.warning("No trial output exists for this invoice.")
        rows_html = "".join(
            f"<tr><td style='padding:4px 12px'>{k}</td><td style='padding:4px 12px'>{v}</td></tr>"
            for k, v in rec["golden"].items()
        )
        st.markdown(
            f"<table><thead><tr><th>Field</th><th>Golden Value</th></tr></thead>"
            f"<tbody>{rows_html}</tbody></table>",
            unsafe_allow_html=True,
        )
        st.subheader("Invoice PDF")
        if pdf_path.exists():
            st.markdown(pdf_iframe(pdf_path), unsafe_allow_html=True)
        else:
            st.info("PDF not available.")
    else:
        label = "✅ All fields match" if rec["status"] == "passed" else "❌ Some fields differ"
        st.subheader(f"Invoice {rec['idx']} — {label}")

        rows          = rec["rows"]
        total_fields  = len(rows)
        matched_fields = sum(1 for r in rows if r["Match"])
        field_accuracy = matched_fields / total_fields * 100 if total_fields else 0

        fc1, fc2, fc3 = st.columns(3)
        fc1.metric("Fields Compared", total_fields)
        fc2.metric("Fields Matched",  matched_fields)
        fc3.metric("Field Accuracy",  f"{field_accuracy:.1f}%")

        show_mismatches_only = st.checkbox("Show mismatches only", value=False)
        display_rows = [r for r in rows if not r["Match"]] if show_mismatches_only else rows

        def row_html(r):
            if r["Match"] and r["Exact"]:
                bg, icon = "#d4edda", "✅"
            elif r["Match"]:
                bg, icon = "#fff9c4", f"🟡 {r['Sim']*100:.0f}%"
            else:
                bg, icon = "#f8d7da", f"❌ {r['Sim']*100:.0f}%"
            return (
                f"<tr style='background:{bg}'>"
                f"<td style='padding:5px 12px'>{r['Field']}</td>"
                f"<td style='padding:5px 12px'>{r['Golden']}</td>"
                f"<td style='padding:5px 12px'>{r['Trial']}</td>"
                f"<td style='padding:5px 12px;text-align:center;white-space:nowrap'>{icon}</td>"
                f"</tr>"
            )

        header = (
            "<thead><tr>"
            "<th style='padding:5px 12px'>Field</th>"
            "<th style='padding:5px 12px'>Golden</th>"
            "<th style='padding:5px 12px'>Trial</th>"
            "<th style='padding:5px 12px'>Match</th>"
            "</tr></thead>"
        )
        body = "".join(row_html(r) for r in display_rows)
        st.markdown(f"<table style='width:100%'>{header}<tbody>{body}</tbody></table>", unsafe_allow_html=True)

        st.subheader("Invoice PDF")
        if pdf_path.exists():
            st.markdown(pdf_iframe(pdf_path), unsafe_allow_html=True)
        else:
            st.info("PDF not available.")

# ── Overview expander ──────────────────────────────────────────────────────────
st.divider()
with st.expander("All invoices overview"):
    def status_badge(status):
        colors = {"passed": "#28a745", "failed": "#dc3545", "not_tested": "#6c757d"}
        labels = {"passed": "Passed", "failed": "Failed", "not_tested": "Not Tested"}
        return f"<span style='color:{colors[status]};font-weight:bold'>{labels[status]}</span>"

    rows_html = ""
    for r in records:
        if r["status"] == "not_tested":
            rows_html += (
                f"<tr><td style='padding:4px 12px'>{r['idx']}</td>"
                f"<td style='padding:4px 12px'>{status_badge('not_tested')}</td>"
                f"<td style='padding:4px 12px'>—</td><td style='padding:4px 12px'>—</td>"
                f"<td style='padding:4px 12px'>—</td></tr>"
            )
        else:
            total   = len(r["rows"])
            matched = sum(1 for row in r["rows"] if row["Match"])
            pct     = f"{matched/total*100:.1f}%" if total else "—"
            rows_html += (
                f"<tr><td style='padding:4px 12px'>{r['idx']}</td>"
                f"<td style='padding:4px 12px'>{status_badge(r['status'])}</td>"
                f"<td style='padding:4px 12px'>{total}</td>"
                f"<td style='padding:4px 12px'>{matched}</td>"
                f"<td style='padding:4px 12px'>{pct}</td></tr>"
            )

    hdr = (
        "<thead><tr>"
        "<th style='padding:4px 12px'>Invoice</th>"
        "<th style='padding:4px 12px'>Status</th>"
        "<th style='padding:4px 12px'>Fields</th>"
        "<th style='padding:4px 12px'>Matched</th>"
        "<th style='padding:4px 12px'>Accuracy</th>"
        "</tr></thead>"
    )
    st.markdown(f"<table style='width:100%'>{hdr}<tbody>{rows_html}</tbody></table>", unsafe_allow_html=True)
